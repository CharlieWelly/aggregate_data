import os
import re

import pandas as pd
from openpyxl import load_workbook

STATEMENT_TYPE = "NHOM_1"
TICKER = "B8"
INTERVAL = "E8"

STATEMENT = {
    "DOUBLE_STATEMENT": "A48",
    "CFLOW_TEST": "A15",
    "Lưu chuyển tiền tệ": {"start": None, "length": None, "first_item": None},
    "Cân đối kế toán": {"start": 14, "length": 122, "first_item": "TÀI SẢN NGẮN HẠN"},
    "Kết quả Kinh doanh": {"start": 14, "length": 25, "first_item": "Doanh số"},
    "Lưu chuyển tiền tệ - Trực tiếp": {
        "start": 14,
        "length": 28,
        "first_item": "Tiền thu từ bán hàng, Cung cấp dịch vụ và DT khác",
    },
    "Lưu chuyển tiền tệ - Gián tiếp": {
        "start": 14,
        "length": 41,
        "first_item": "Lãi trước thuế",
    },
    "Lưu chuyển tiền tệ - Gián tiếp - 2": {
        "start": 47,
        "length": 41,
        "first_item": "Lãi trước thuế",
    },
    "Thuyết minh": {"start": 14, "length": 157, "first_item": "Tiền"},
}


def setUp(statement_type):
    import setting

    global STATEMENT_TYPE, TICKER, INTERVAL, STATEMENT

    STATEMENT_TYPE = statement_type
    TICKER = setting.TICKER
    INTERVAL = setting.INTERVAL
    STATEMENT = setting.STATEMENT_TYPE[statement_type]


class WorkBook(object):
    def __init__(self, file):
        self.wb = load_workbook(file)
        self.sheetnames = self.wb.sheetnames
        assert len(self.sheetnames) == 4
        self.statement_type = STATEMENT_TYPE

    def __getitem__(self, key):
        return self.wb[key]

    def contain_double_statement(self):
        double_statement = STATEMENT["DOUBLE_STATEMENT"]
        val = self.wb["Lưu chuyển tiền tệ"][double_statement].value
        return (
            val
            and val.strip()
            == STATEMENT["Lưu chuyển tiền tệ - Gián tiếp - 2"]["first_item"]
        )

    def export_sheets(self):
        sheets = []
        for sheetname in self.sheetnames:
            if not sheetname == "Lưu chuyển tiền tệ":
                sheets.append(WorkSheet(self, sheetname))
            else:
                sheets.append(CashFlowSheet(self, sheetname))
                if self.contain_double_statement():
                    sheets.append(CashFlowSheet2(self, sheetname))
        return sheets


class WorkSheet(object):
    def __init__(self, wb, sheetname):
        self.wb = wb
        self.sheetname = sheetname
        self.content = self.wb[self.sheetname]
        self.statement_type = self.wb.statement_type
        self.start = STATEMENT[self.sheetname]["start"]
        self.length = STATEMENT[self.sheetname]["length"]
        self.first_item = STATEMENT[self.sheetname]["first_item"]
        self.ticker = self.content[TICKER].value.strip()
        self.interval = self.content[INTERVAL].value.strip()

    def __str__(self):
        return "<Worksheet: %s>" % self.sheetname

    def __repr__(self):
        return "<Worksheet: %s>" % self.sheetname

    def to_df(self):
        df = pd.DataFrame(self.content.values)
        df.iloc[10, 0] = "items"
        df.columns = df.iloc[10]
        df = df.iloc[self.start : self.start + self.length]
        df.dropna(axis="columns", how="all", inplace=True)
        df.dropna(axis="index", how="all", inplace=True)
        df.dropna(subset=["items"], inplace=True)
        df["items"] = df["items"].str.strip()
        df.reset_index(drop=True, inplace=True)

        if len(df.columns) > 1:
            self.drop_nodatetimecol(df)
            self.validate_design(df)
            df = pd.melt(df, id_vars=["items"], var_name="period", ignore_index=False)
            df["statement_name"] = self.sheetname
            df["statement_interval"] = self.interval
            df["business_type"] = self.statement_type
            df["ticker"] = self.ticker
            return df
        else:
            return pd.DataFrame()

    def drop_nodatetimecol(self, df):
        pattern = re.compile(r"\s*(?:Q\d\s*/)?\s*\d{4}\s*")
        for col in df.columns[1:]:
            if not pattern.match(col):
                df.drop(columns=col, inplace=True)

    def validate_design(self, df):
        try:
            first_item = df.loc[0, "items"]
            assert first_item == self.first_item
        except AssertionError as e:
            print(f"{self.ticker} {self.sheetname} has invalid design")
            print(f"expect {self.first_item}, but receive {first_item}")
            raise e


class CashFlowSheet(WorkSheet):
    def __init__(self, wb, sheetname):
        WorkSheet.__init__(self, wb, sheetname)

        if self.contain_direct_cashflow():
            self.sheetname = "Lưu chuyển tiền tệ - Trực tiếp"
        else:
            self.sheetname = "Lưu chuyển tiền tệ - Gián tiếp"

        self.start = STATEMENT[self.sheetname]["start"]
        self.length = STATEMENT[self.sheetname]["length"]
        self.first_item = STATEMENT[self.sheetname]["first_item"]

    def contain_direct_cashflow(self):
        cflow_test = STATEMENT["CFLOW_TEST"]
        val = self.content[cflow_test].value
        return (
            val
            and val.strip() == STATEMENT["Lưu chuyển tiền tệ - Trực tiếp"]["first_item"]
        )


class CashFlowSheet2(WorkSheet):
    def __init__(self, wb, sheetname):
        WorkSheet.__init__(self, wb, sheetname)
        self.sheetname = "Lưu chuyển tiền tệ - Gián tiếp"
        self.subfix = " - 2"
        self.start = STATEMENT[self.sheetname + self.subfix]["start"]
        self.length = STATEMENT[self.sheetname + self.subfix]["length"]
        self.first_item = STATEMENT[self.sheetname + self.subfix]["first_item"]


def loop_thruBook(wb):
    return pd.concat([ws.to_df() for ws in wb.export_sheets()])


def loop_thruFile(excel_file):
    wb = WorkBook(excel_file)
    return pd.concat([ws.to_df() for ws in wb.export_sheets()])


def loop_thruFolder(folder):
    files = os.listdir(folder)
    pat = re.compile(r".*FiinPro_BCTC_.*")
    df = pd.concat(
        [loop_thruFile(f"{folder}/{file}") for file in files if pat.match(file)]
    )
    return df


def toUserFriendly(df):
    df = (
        df.groupby(["ticker", "period", "statement_name", "items"], sort=False)["value"]
        .sum()
        .unstack(["statement_name", "items"])
    )
    return df
